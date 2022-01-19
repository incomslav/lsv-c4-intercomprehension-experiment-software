from collections import defaultdict

from Common.constants import *
from Common.Enums import *
from Users.enums import *

from django.shortcuts import render
from django.views.generic import View
from django.http import HttpResponseRedirect
from django.http import HttpResponse
# from django_datatables_view.base_datatable_view import BaseDatatableView
# from django.core.files.storage import FileSystemStorage

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.write_only import WriteOnlyCell
from openpyxl.writer.excel import save_virtual_workbook

from FreeTranslationExperiment.models import *
import json


def getHesitationAndTypingTimeFromResponseChanges(json_data):
    """
    get hesitation and typing time from response changes(i.e list of time stamps)
    """
    if json_data != None:
        changes = json.loads(json_data)
    else:
        changes = []
    out = []
    timestamps = []
    states = []

    if len(changes) >= 2:
        hesitation_time = float(changes[1][1]) - float(changes[0][1])
    else:
        # hesitation_time = float("inf")
        hesitation_time = 0

    if len(changes) >= 3:
        time_spent_typing = float(changes[-1][1]) - float(changes[0][1])
    else:
        # time_spent_typing = float("inf")
        time_spent_typing = 0

    return hesitation_time, time_spent_typing


def getQuestionsByExperimentId(experimentId, filters, forExcel):
    """
    get all question's answer count by experiment id
    :param experimentId:
    :return:
    """
    data = []
    common = {}

    filter_ranks = filters["ranks"]
    from_date = str(filters["fromDate"]).strip()
    to_date = str(filters["toDate"]).strip()

    isStimulusLanguageNotKnown = filters["isStimulusLanguageNotKnown"]

    experiment = Experiment.objects.filter(id=experimentId).select_subclasses()[0]

    # experiment = Experiment.objects.get(id=experimentId)
    foreign_language = str(experiment.foreign_language)

    common["ExperimentId"] = experiment.id
    common["ExperimentName"] = experiment.experiment_name
    sr = 0
    # get all participations in this experiment
    for question in experiment.experiment_questions.select_subclasses():

        sr += 1
        # question.
        num_answers = 0
        num_no_answers = 0
        total_answers = 0
        abs_correct_answers = 0
        normalized_correct_answers = 0
        incorrect_answers = 0

        abs_correct_time = {'total': 0, 'init_hesitation': 0, 'typing': 0, 'submit_hesitation': 0}
        norm_correct_time = {'total': 0, 'init_hesitation': 0, 'typing': 0, 'submit_hesitation': 0}
        incorrect_time = {'total': 0, 'init_hesitation': 0, 'typing': 0, 'submit_hesitation': 0}
        no_ans_time = {'total': 0, 'init_hesitation': 0, 'typing': 0, 'submit_hesitation': 0}

        if from_date != '' and to_date != '':
            UserAnswers = UserAnswer.objects.filter(answered_question=question,
                                                    answer_date__range=[from_date, to_date]).select_subclasses()
        else:
            UserAnswers = UserAnswer.objects.filter(answered_question=question).select_subclasses()

        for user_answer in UserAnswers:

            ans_rank = user_answer.experiment_participation.getAnswerRank()
            user_languages = [str(language) for language in user_answer.answering_user.getAllLanguages()]

            userKnowForeignLanguage = foreign_language in user_languages

            if userKnowForeignLanguage == True and isStimulusLanguageNotKnown == '1':
                userKnowForeignLanguage = False
            else:
                userKnowForeignLanguage = True

            # filter for ans_rank ,if ans_rank is in selected ranks
            if ans_rank in filter_ranks and userKnowForeignLanguage == True:

                total_answers += 1

                if user_answer.answer_given:
                    num_answers += 1
                    if user_answer.isExactlyCorrectAnswer():
                        abs_correct_answers += 1
                        if user_answer.time_spent:
                            abs_correct_time_spent = user_answer.time_spent
                        else:
                            abs_correct_time_spent = 0

                        abs_correct_initial_hesitation_time, abs_correct_typing_time = getHesitationAndTypingTimeFromResponseChanges(
                            user_answer.result_changes)
                        abs_correct_submission_hesitation_time = abs_correct_time_spent - abs_correct_typing_time

                        abs_correct_time['total'] += abs_correct_time_spent
                        abs_correct_time['init_hesitation'] += abs_correct_initial_hesitation_time
                        abs_correct_time['typing'] += abs_correct_typing_time
                        abs_correct_time['submit_hesitation'] += abs_correct_submission_hesitation_time

                    elif user_answer.isCorrectAnswer():
                        normalized_correct_answers += 1

                        if user_answer.time_spent:
                            norm_correct_time_spent = user_answer.time_spent
                        else:
                            norm_correct_time_spent = 0
                        norm_correct_initial_hesitation_time, norm_correct_typing_time = getHesitationAndTypingTimeFromResponseChanges(
                            user_answer.result_changes)
                        norm_correct_submission_hesitation_time = norm_correct_time_spent - norm_correct_typing_time

                        norm_correct_time['total'] += norm_correct_time_spent
                        norm_correct_time['init_hesitation'] += norm_correct_initial_hesitation_time
                        norm_correct_time['typing'] += norm_correct_typing_time
                        norm_correct_time['submit_hesitation'] += norm_correct_submission_hesitation_time

                    else:
                        incorrect_answers += 1

                        if user_answer.time_spent:
                            incorrect_time_spent = user_answer.time_spent
                        else:
                            incorrect_time_spent = 0
                        incorrect_initial_hesitation_time, incorrect_typing_time = getHesitationAndTypingTimeFromResponseChanges(
                            user_answer.result_changes)
                        incorrect_submission_hesitation_time = incorrect_time_spent - incorrect_typing_time

                        incorrect_time['total'] += incorrect_time_spent
                        incorrect_time['init_hesitation'] += incorrect_initial_hesitation_time
                        incorrect_time['typing'] += incorrect_typing_time
                        incorrect_time['submit_hesitation'] += incorrect_submission_hesitation_time

                else:
                    num_no_answers += 1

                    if user_answer.time_spent:
                        no_ans_time_spent = user_answer.time_spent
                    else:
                        no_ans_time_spent = 0
                    no_ans_initial_hesitation_time, no_ans_typing_time = getHesitationAndTypingTimeFromResponseChanges(
                        user_answer.result_changes)
                    no_ans_submission_hesitation_time = no_ans_time_spent - no_ans_typing_time

                    no_ans_time['total'] += no_ans_time_spent
                    no_ans_time['init_hesitation'] += no_ans_initial_hesitation_time
                    no_ans_time['typing'] += no_ans_typing_time
                    no_ans_time['submit_hesitation'] += no_ans_submission_hesitation_time
                    # end of change

        if total_answers > 0:
            perc_abs_correct = round(abs_correct_answers / total_answers, 3)
            perc_normalized_correct = round(normalized_correct_answers / total_answers, 3)
            perc_incorrect = round(incorrect_answers / total_answers, 3)

            perc_no_answers = round(num_no_answers / total_answers, 3)

            total_time_spent = abs_correct_time['total'] + norm_correct_time['total'] + incorrect_time['total'] + \
                               no_ans_time['total']

            total_initial_hesitation_time = abs_correct_time['init_hesitation'] + norm_correct_time['init_hesitation'] + \
                                            incorrect_time['init_hesitation'] + \
                                            no_ans_time['init_hesitation']

            total_typing_time = abs_correct_time['typing'] + norm_correct_time['typing'] + incorrect_time['typing'] + \
                                no_ans_time['typing']

            total_submission_hesitation_time = abs_correct_time['submit_hesitation'] + norm_correct_time[
                'submit_hesitation'] + incorrect_time['submit_hesitation'] + \
                                               no_ans_time['submit_hesitation']

            avg_total_time_spent = round((total_time_spent / total_answers) / 1000, 2)
            avg_initial_hesitation_time = round((total_initial_hesitation_time / total_answers) / 1000, 2)
            avg_typing_time = round((total_typing_time / total_answers) / 1000, 2)
            avg_submission_hesitation_time = round((total_submission_hesitation_time / total_answers) / 1000, 2)

        else:
            perc_abs_correct = "N/A"
            perc_normalized_correct = "N/A"
            perc_incorrect = "N/A"
            perc_no_answers = "N/A"

            avg_total_time_spent = "N/A"
            avg_initial_hesitation_time = "N/A"
            avg_typing_time = "N/A"
            avg_submission_hesitation_time = "N/A"

        if abs_correct_answers > 0:
            abs_avg_total_time_spent = round((abs_correct_time['total'] / abs_correct_answers) / 1000, 2)
            abs_avg_initial_hesitation_time = round((abs_correct_time['init_hesitation'] / abs_correct_answers) / 1000,
                                                    2)
            abs_avg_typing_time = round((abs_correct_time['typing'] / abs_correct_answers) / 1000, 2)
            abs_avg_submit_hesitation_time = round((abs_correct_time['submit_hesitation'] / abs_correct_answers) / 1000,
                                                   2)
        else:
            abs_avg_total_time_spent = "N/A"
            abs_avg_initial_hesitation_time = "N/A"
            abs_avg_typing_time = "N/A"
            abs_avg_submit_hesitation_time = "N/A"
        if normalized_correct_answers > 0:
            norm_avg_total_time_spent = round((norm_correct_time['total'] / normalized_correct_answers) / 1000, 2)
            norm_avg_initial_hesitation_time = round(
                (norm_correct_time['init_hesitation'] / normalized_correct_answers) / 1000, 2)
            norm_avg_typing_time = round((norm_correct_time['typing'] / normalized_correct_answers) / 1000, 2)
            norm_avg_submit_hesitation_time = round(
                (norm_correct_time['submit_hesitation'] / normalized_correct_answers) / 1000, 2)
        else:
            norm_avg_total_time_spent = "N/A"
            norm_avg_initial_hesitation_time = "N/A"
            norm_avg_typing_time = "N/A"
            norm_avg_submit_hesitation_time = "N/A"

        if incorrect_answers > 0:
            incorrect_avg_total_time_spent = round((incorrect_time['total'] / incorrect_answers) / 1000, 2)
            incorrect_avg_initial_hesitation_time = round(
                (incorrect_time['init_hesitation'] / incorrect_answers) / 1000, 2)
            incorrect_avg_typing_time = round((incorrect_time['typing'] / incorrect_answers) / 1000, 2)
            incorrect_avg_submit_hesitation_time = round(
                (incorrect_time['submit_hesitation'] / incorrect_answers) / 1000, 2)
        else:
            incorrect_avg_total_time_spent = "N/A"
            incorrect_avg_initial_hesitation_time = "N/A"
            incorrect_avg_typing_time = "N/A"
            incorrect_avg_submit_hesitation_time = "N/A"

        if num_no_answers > 0:
            no_ans_avg_total_time_spent = round((no_ans_time['total'] / num_no_answers) / 1000, 2)
            no_ans_avg_initial_hesitation_time = round(
                (no_ans_time['init_hesitation'] / num_no_answers) / 1000, 2)
            no_ans_avg_typing_time = round((no_ans_time['typing'] / num_no_answers) / 1000, 2)
            no_ans_avg_submit_hesitation_time = round(
                (no_ans_time['submit_hesitation'] / num_no_answers) / 1000, 2)
        else:
            no_ans_avg_total_time_spent = "N/A"
            no_ans_avg_initial_hesitation_time = "N/A"
            no_ans_avg_typing_time = "N/A"
            no_ans_avg_submit_hesitation_time = "N/A"

        if not forExcel:
            url = "'admin/ExportUserAnswersPreview/{}/{}'".format(experiment.id, question.id)
            question_word = '<a onclick="redirectTo({})">{}</a>'.format(url, str(question))
        else:
            question_word = str(question)

        data.append([sr, question.id, question_word, total_answers, abs_correct_answers, normalized_correct_answers,
                     incorrect_answers, num_no_answers, perc_abs_correct, perc_normalized_correct, perc_incorrect,
                     perc_no_answers, avg_total_time_spent, avg_initial_hesitation_time, avg_typing_time,
                     avg_submission_hesitation_time, abs_avg_total_time_spent, abs_avg_initial_hesitation_time,
                     abs_avg_typing_time, abs_avg_submit_hesitation_time,
                     norm_avg_total_time_spent, norm_avg_initial_hesitation_time, norm_avg_typing_time,
                     norm_avg_submit_hesitation_time,
                     incorrect_avg_total_time_spent, incorrect_avg_initial_hesitation_time, incorrect_avg_typing_time,
                     incorrect_avg_submit_hesitation_time,
                     no_ans_avg_total_time_spent, no_ans_avg_initial_hesitation_time, no_ans_avg_typing_time,
                     no_ans_avg_submit_hesitation_time])

    return common, data


def getUserAnswersByExperimentAndQuestionId(experimentId, questionId, forExcel):
    """
    get all user's answers by experiment and question id
    :param experimentId:
    :param questionId:
    """
    data = []
    common = {}

    experiment = Experiment.objects.get(id=experimentId)
    # get question in this experiment
    question = experiment.experiment_questions.get_subclass(id=questionId)
    common["ExperimentId"] = experiment.id
    common["ExperimentName"] = experiment.experiment_name
    sr = 0
    common["Question"] = str(question)

    for user_answer in UserAnswer.objects.filter(answered_question=question).select_subclasses():
        sr += 1

        user = user_answer.answering_user
        user_id = user.id
        if user_answer.answer_given:
            answer_given = user_answer.native_word
            abs_correct = user_answer.isExactlyCorrectAnswer()
            norm_correct = user_answer.isCorrectAnswer()
            incorrect = not (abs_correct or norm_correct)
            missing_ans = False
        else:
            answer_given = "N/A"
            abs_correct = "N/A"
            norm_correct = "N/A"
            incorrect = "N/A"
            missing_ans = True

        gender = user.gender

        if user_answer.time_spent:
            total_time = user_answer.time_spent
        else:
            total_time = 0

        initial_hesitation_time, typing_time = getHesitationAndTypingTimeFromResponseChanges(user_answer.result_changes)
        submission_hesitation_time = total_time - typing_time
        age = user.age
        living_location = user.location
        living_location_duration = user.location_living_duration

        education_countries = user.getEducationCountryStays()
        language_skills = user.getAllLanguageSkills()

        if not forExcel:

            lsLanguageSkills = []
            lsLanguageSkills.append("<table><tr>")
            languages = ''.join(["<th colspan='8'>{}</th>".format(l.language) for l in language_skills])
            lsLanguageSkills.append(languages)
            lsLanguageSkills.append("</tr><tr>")
            for l in language_skills:
                lsLanguageSkills.append(
                    "<td>Reading</td><td>Writing</td><td>Listening</td><td>Speaking</td><td>Native</td><td>Home</td><td>LearnedFor</td><td>LivingExposure</td>")

            lsLanguageSkills.append("</tr><tr>")

            skill_levels = ''.join([
                "<td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td><td>{}</td>".format(
                    l.reading_level, l.writing_level, l.listening_level, l.speaking_level, l.is_native_language,
                    l.used_at_home, l.exposure_through_learning, l.exposure_through_living) for l in language_skills])
            lsLanguageSkills.append(skill_levels)

            lsLanguageSkills.append("</tr></table>")
            lsLanguageSkills = ''.join(lsLanguageSkills)

            lsEducationCountries = []

            lsEducationCountries.append("<table><tr>")
            ed_countries = ''.join(["<th>{}</th>".format(ec.country.country_name) for ec in education_countries])
            lsEducationCountries.append(ed_countries)
            lsEducationCountries.append("</tr><tr>")
            ed_duration = ''.join(["<td>{}</td>".format(ec.duration) for ec in education_countries])
            lsEducationCountries.append(ed_duration)
            lsEducationCountries.append("</tr></table>")
            lsEducationCountries = ''.join(lsEducationCountries)
        else:

            lsEducationCountries = education_countries
            lsLanguageSkills = language_skills

        data.append(
            [sr, user_id, str(answer_given), str(abs_correct), str(norm_correct), str(incorrect), str(missing_ans),
             str(round(total_time / 1000, 2)), str(round(initial_hesitation_time / 1000, 2)),
             str(round(typing_time / 1000, 2)),
             str(round(submission_hesitation_time / 1000, 2)), str(age), str(gender), str(living_location),
             str(living_location_duration), lsEducationCountries, lsLanguageSkills])

    return common, data


class ExportPreview(View):
    def get(self, request, exp_id):

        all_ranks = Experiment.objects.get(id=exp_id).getNumbersOfCollectedAnswers()
        all_ranks = [x + 1 for x in all_ranks.keys()]
        if "export_preview_filters" in request.session:
            filters = request.session["export_preview_filters"]
            common, data = getQuestionsByExperimentId(int(exp_id), filters, forExcel=False)
        else:

            filters = {}
            ranks = [int(r) - 1 for r in all_ranks]
            ranks.append(None)
            filters["ranks"] = ranks
            filters["fromDate"] = ""
            filters["toDate"] = ""
            filters["isStimulusLanguageNotKnown"] = '0'

            common, data = getQuestionsByExperimentId(int(exp_id), filters, forExcel=False)

        return render(request, "adminpanel/ExportPreview.html",
                      {'common_data': common, 'data': data, 'ranks': all_ranks, 'filter': filters})

    def post(self, request, exp_id):
        try:
            all_ranks = Experiment.objects.get(id=exp_id).getNumbersOfCollectedAnswers()
            all_ranks = [x + 1 for x in all_ranks.keys()]
            all_ranks_zero_index = [int(r) - 1 for r in all_ranks]
            all_ranks_zero_index.append(None)
            filters = {}
            if "export_preview_filters" in request.session:
                filters = request.session["export_preview_filters"]
            else:
                filters["ranks"] = all_ranks_zero_index
                filters["fromDate"] = ""
                filters["toDate"] = ""
                filters["isStimulusLanguageNotKnown"] = 0

            isFilter = request.POST['hdnFilter']

            # set filters
            if isFilter == '1':
                filter_ranks = request.POST.getlist('chk_rank')
                from_date = request.POST['txtFromDate']
                to_date = request.POST['txtToDate']
                is_stimulus_language_not_known = request.POST['hdn_is_not_know_stimulus_language']

                if filter_ranks == []:
                    filter_ranks = all_ranks_zero_index
                else:
                    filter_ranks = [int(r) - 1 for r in filter_ranks]

                filters["ranks"] = filter_ranks
                filters["fromDate"] = from_date
                filters["toDate"] = to_date
                filters["isStimulusLanguageNotKnown"] = is_stimulus_language_not_known

                experimentData, questionData = getQuestionsByExperimentId(int(exp_id), filters, forExcel=True)
                request.session["export_preview_filters"] = filters
                return render(request, "adminpanel/ExportPreview.html",
                              {'common_data': experimentData, 'data': questionData, 'ranks': all_ranks,
                               'filter': filters})
            # remove filter
            elif isFilter == '2':
                if "export_preview_filters" in request.session:
                    del request.session["export_preview_filters"]
                return HttpResponseRedirect('/en/admin/ExportPreview/{}'.format(exp_id))


            # download excel file
            experimentData, questionData = getQuestionsByExperimentId(int(exp_id), filters, forExcel=True)

            wb = Workbook(write_only=True)
            ws = wb.create_sheet()

            # header
            cell_questionId = WriteOnlyCell(ws, value="Question ID")
            cell_questionId.font = Font(bold=True)
            cell_questionWord = WriteOnlyCell(ws, value="Stimulus")
            cell_questionWord.font = Font(bold=True)
            cell_totalAnswers = WriteOnlyCell(ws, value="Total Answers")
            cell_totalAnswers.font = Font(bold=True)
            cell_abs_correct = WriteOnlyCell(ws, value="# Absolutely Correct")
            cell_abs_correct.font = Font(bold=True)
            cell_normalized_correct = WriteOnlyCell(ws, value="# Normalized Correct")
            cell_normalized_correct.font = Font(bold=True)
            cell_incorrect = WriteOnlyCell(ws, value="# Incorrect")
            cell_incorrect.font = Font(bold=True)
            cell_no_answer = WriteOnlyCell(ws, value="# No Answer")
            cell_no_answer.font = Font(bold=True)
            cell_p_abs_correct = WriteOnlyCell(ws, value="% Absolutely Correct")
            cell_p_abs_correct.font = Font(bold=True)
            cell_p_norm_correct = WriteOnlyCell(ws, value="% Normalized Correct")
            cell_p_norm_correct.font = Font(bold=True)
            cell_p_incorrect = WriteOnlyCell(ws, value="% Incorrect")
            cell_p_incorrect.font = Font(bold=True)
            cell_p_no_ans = WriteOnlyCell(ws, value="% No Answer")
            cell_p_no_ans.font = Font(bold=True)

            # Columns for Times
            cell_avg_total_time = WriteOnlyCell(ws, value="Avg Total Time")
            cell_avg_total_time.font = Font(bold=True)
            cell_avg_init_hesitation_time = WriteOnlyCell(ws, value="Avg Initial Hesitation Time")
            cell_avg_init_hesitation_time.font = Font(bold=True)
            cell_avg_typing_time = WriteOnlyCell(ws, value="Avg Typing Time")
            cell_avg_typing_time.font = Font(bold=True)
            cell_avg_submit_hesitation_time = WriteOnlyCell(ws, value="Avg Submission Hesitation Time")
            cell_avg_submit_hesitation_time.font = Font(bold=True)

            # for absolute correct answers time
            cell_abs_avg_total_time = WriteOnlyCell(ws, value="Absolutely Correct Avg Total Time")
            cell_abs_avg_total_time.font = Font(bold=True)
            cell_abs_avg_init_hesitation_time = WriteOnlyCell(ws,
                                                              value="Absolutely Correct Avg Initial Hesitation Time")
            cell_abs_avg_init_hesitation_time.font = Font(bold=True)
            cell_abs_avg_typing_time = WriteOnlyCell(ws, value="Absolutely Correct Avg Typing Time")
            cell_abs_avg_typing_time.font = Font(bold=True)
            cell_abs_avg_submit_hesitation_time = WriteOnlyCell(ws,
                                                                value="Absolutely Correct Avg Submission Hesitation Time")
            cell_abs_avg_submit_hesitation_time.font = Font(bold=True)

            # for normalized correct answers time
            cell_norm_avg_total_time = WriteOnlyCell(ws, value="Normalized Correct Avg Total Time")
            cell_norm_avg_total_time.font = Font(bold=True)
            cell_norm_avg_init_hesitation_time = WriteOnlyCell(ws,
                                                               value="Normalized Correct Avg Initial Hesitation Time")
            cell_norm_avg_init_hesitation_time.font = Font(bold=True)
            cell_norm_avg_typing_time = WriteOnlyCell(ws, value="Normalized Correct Avg Typing Time")
            cell_norm_avg_typing_time.font = Font(bold=True)
            cell_norm_avg_submit_hesitation_time = WriteOnlyCell(ws,
                                                                 value="Normalized Correct Avg Submission Hesitation Time")
            cell_norm_avg_submit_hesitation_time.font = Font(bold=True)

            # for incorrect answers time
            cell_incorrect_avg_total_time = WriteOnlyCell(ws, value="Incorrect Avg Total Time")
            cell_incorrect_avg_total_time.font = Font(bold=True)
            cell_incorrect_avg_init_hesitation_time = WriteOnlyCell(ws, value="Incorrect Avg Initial Hesitation Time")
            cell_incorrect_avg_init_hesitation_time.font = Font(bold=True)
            cell_incorrect_avg_typing_time = WriteOnlyCell(ws, value="Incorrect Avg Typing Time")
            cell_incorrect_avg_typing_time.font = Font(bold=True)
            cell_incorrect_avg_submit_hesitation_time = WriteOnlyCell(ws,
                                                                      value="Incorrect Avg Submission Hesitation Time")
            cell_incorrect_avg_submit_hesitation_time.font = Font(bold=True)

            # for no answers time
            cell_missing_avg_total_time = WriteOnlyCell(ws, value="Missing Answers Avg Total Time")
            cell_missing_avg_total_time.font = Font(bold=True)
            cell_missing_avg_init_hesitation_time = WriteOnlyCell(ws,
                                                                  value="Missing Answers Avg Initial Hesitation Time")
            cell_missing_avg_init_hesitation_time.font = Font(bold=True)
            cell_missing_avg_typing_time = WriteOnlyCell(ws, value="Missing Answers Avg Typing Time")
            cell_missing_avg_typing_time.font = Font(bold=True)
            cell_missing_avg_submit_hesitation_time = WriteOnlyCell(ws,
                                                                    value="Missing Answers Avg Submission Hesitation Time")
            cell_missing_avg_submit_hesitation_time.font = Font(bold=True)

            ws.append([cell_questionId, cell_questionWord, cell_totalAnswers, cell_abs_correct,
                       cell_normalized_correct, cell_incorrect, cell_no_answer, cell_p_abs_correct, cell_p_norm_correct,
                       cell_p_incorrect, cell_p_no_ans, cell_avg_total_time, cell_avg_init_hesitation_time,
                       cell_avg_typing_time, cell_avg_submit_hesitation_time,
                       cell_abs_avg_total_time, cell_abs_avg_init_hesitation_time, cell_abs_avg_typing_time,
                       cell_abs_avg_submit_hesitation_time,
                       cell_norm_avg_total_time, cell_norm_avg_init_hesitation_time, cell_norm_avg_typing_time,
                       cell_norm_avg_submit_hesitation_time,
                       cell_incorrect_avg_total_time, cell_incorrect_avg_init_hesitation_time,
                       cell_incorrect_avg_typing_time, cell_incorrect_avg_submit_hesitation_time,
                       cell_missing_avg_total_time, cell_missing_avg_init_hesitation_time, cell_missing_avg_typing_time,
                       cell_missing_avg_submit_hesitation_time])

            for d in questionData:
                # ignoring sr column
                ws.append(d[1:])

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={}_questions.xlsx'.format(
                experimentData['ExperimentName'].replace(' ', '_'))

            wb.save(response)
            return response

        except Exception as e:
            return HttpResponse(str(e))


class ExportUserAnswersPreview(View):
    def get(self, request, exp_id, q_id):

        common, data = getUserAnswersByExperimentAndQuestionId(int(exp_id), int(q_id), forExcel=False)
        return render(request, "adminpanel/ExportUserAnswersPreview.html", {'common_data': common, 'data': data})

    def post(self, request, exp_id, q_id):
        try:

            experimentData, questionData = getUserAnswersByExperimentAndQuestionId(int(exp_id), int(q_id),
                                                                                   forExcel=True)

            wb = Workbook()
            ws = wb.get_active_sheet()

            # header
            ws.cell(row=1, column=1).value = 'User Id'
            ws.cell(row=1, column=2).value = 'Given Answer'
            ws.cell(row=1, column=3).value = 'Absolutely Correct'
            ws.cell(row=1, column=4).value = 'Normalized Correct'
            ws.cell(row=1, column=5).value = 'Incorrect'
            ws.cell(row=1, column=6).value = 'No Answer'
            ws.cell(row=1, column=7).value = 'Total Time'
            ws.cell(row=1, column=8).value = 'Initial Hesitation Time'
            ws.cell(row=1, column=9).value = 'Typing Time'
            ws.cell(row=1, column=10).value = 'Submission Hesitation Time'
            ws.cell(row=1, column=11).value = 'Age'
            ws.cell(row=1, column=12).value = 'Gender'
            ws.cell(row=1, column=13).value = 'Living Location'
            ws.cell(row=1, column=14).value = 'Living Period'
            ws.cell(row=1, column=15).value = 'Education Countries'
            r = 2
            # find number of maximum education countries
            maximum_ed_countries = len(max([ec for ec in [d[15] for d in questionData]], key=len))
            ws.merge_cells(start_row=1, start_column=15, end_row=1, end_column=15 + maximum_ed_countries - 1)
            ws.cell(row=1, column=15 + maximum_ed_countries).value = 'Language Skills'

            for i in range(1, 15 + maximum_ed_countries + 1):
                ws.cell(row=1, column=i).font = Font(bold=True)

            for data in questionData:

                ws.cell(row=r, column=1).value = data[1]
                ws.cell(row=r, column=2).value = data[2]
                ws.cell(row=r, column=3).value = data[3]
                ws.cell(row=r, column=4).value = data[4]
                ws.cell(row=r, column=5).value = data[5]
                ws.cell(row=r, column=6).value = data[6]
                ws.cell(row=r, column=7).value = data[7]
                ws.cell(row=r, column=8).value = data[8]
                ws.cell(row=r, column=9).value = data[9]
                ws.cell(row=r, column=10).value = data[10]
                ws.cell(row=r, column=11).value = data[11]
                ws.cell(row=r, column=12).value = data[12]
                ws.cell(row=r, column=13).value = data[13]
                ws.cell(row=r, column=14).value = data[14]
                c = 15

                ec = 0
                for education_countries in data[15]:
                    ec += 1
                    ws.cell(row=r, column=c).value = str(education_countries.country.country_name)
                    ws.cell(row=r + 1, column=c).value = str(education_countries.duration)
                    c += 1
                if ec != maximum_ed_countries:
                    c = c + maximum_ed_countries - 1

                for user_language_skills in data[16]:
                    ws.cell(row=r, column=c).value = str(user_language_skills.language)
                    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 7)

                    ws.cell(row=r + 1, column=c).value = 'Reading Skill'
                    ws.cell(row=r + 1, column=c + 1).value = 'Writing Skill'
                    ws.cell(row=r + 1, column=c + 2).value = 'Listening Skill'
                    ws.cell(row=r + 1, column=c + 3).value = 'Speaking Skill'
                    ws.cell(row=r + 1, column=c + 4).value = 'Is Native?'
                    ws.cell(row=r + 1, column=c + 5).value = 'Is Home Language?'
                    ws.cell(row=r + 1, column=c + 6).value = 'Learning Period'
                    ws.cell(row=r + 1, column=c + 7).value = 'Exposure Through Living'

                    ws.cell(row=r + 2, column=c).value = str(user_language_skills.reading_level)
                    ws.cell(row=r + 2, column=c + 1).value = str(user_language_skills.writing_level)
                    ws.cell(row=r + 2, column=c + 2).value = str(user_language_skills.listening_level)
                    ws.cell(row=r + 2, column=c + 3).value = str(user_language_skills.speaking_level)
                    ws.cell(row=r + 2, column=c + 4).value = str(user_language_skills.is_native_language)
                    ws.cell(row=r + 2, column=c + 5).value = str(user_language_skills.used_at_home)
                    ws.cell(row=r + 2, column=c + 6).value = str(user_language_skills.exposure_through_learning)
                    ws.cell(row=r + 2, column=c + 7).value = str(user_language_skills.exposure_through_living)
                    c += 8

                r += 3

            response = HttpResponse(content=save_virtual_workbook(wb),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={}_user_answers_data.xlsx'.format(
                experimentData['ExperimentName'].replace(' ', '_'))
            return response

        except Exception as e:
            return HttpResponse(str(e))
